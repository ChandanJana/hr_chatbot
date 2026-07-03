"""Helper to read documents from text files.

This module provides functions to read single or multiple text files.
You can add PDF support here later.
"""

from pathlib import Path
from typing import Dict

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def read_text_file(path: str) -> str:
    """Read a UTF-8 text file and return its contents.

    Args:
        path: Path to the `.txt` file to read.

    Returns:
        The full file contents as a string.
    """
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def read_pdf_file(path: str) -> str:
    """Extract text from a PDF file.

    Args:
        path: Path to the `.pdf` file to read.

    Returns:
        The extracted text content from all PDF pages, joined by newlines.

    Raises:
        RuntimeError: If PyPDF2 is not installed.
    """

    # In Python, None means:
    # No value or Nothing. It is often used as a default value for function
    # arguments or to indicate the absence of a return value.
    # if PyPDF2 is None: is a Python pattern used to check whether a 
    # library was imported successfully or not.
    if PyPDF2 is None:
        raise RuntimeError(
            "PyPDF2 is required to read PDF files. Install with `pip install PyPDF2`."
        )

    text = []
    """
    This statement is used to open a file in binary read mode and automatically close it when the block finishes.
    The 'rb' mode in open(path, "rb") stands for "read binary".
    The 'with' statement automatically manages the file resource.
    with open(path, "rb") as f:
    ...
    is equivalent to:
    f = open(path, "rb")
    try:
    ...
    finally:
     f.close()
    You do not need to call:
    f.close() 
    explicitly; it will be closed automatically when the block is exited.
    """

    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text() or ""
            text.append(page_text)

    return "\n".join(text)


def read_docx_file(path: str) -> str:
    """Extract text from a Word document (.docx) file.

    Args:
        path: Path to the `.docx` file to read.

    Returns:
        The extracted text content from all paragraphs in the document,
        joined by newlines.

    Raises:
        RuntimeError: If python-docx is not installed.
    """
    if Document is None:
        raise RuntimeError(
            "python-docx is required to read DOCX files. Install with `pip install python-docx`."
        )

    text = []
    doc = Document(path)
    for paragraph in doc.paragraphs:
        if paragraph.text.strip():
            text.append(paragraph.text)

    return "\n".join(text)


def read_excel_file(path: str) -> str:
    """Extract text from an Excel file (.xlsx or .xls).

    Args:
        path: Path to the Excel file to read.

    Returns:
        The extracted text content from all cells, with rows separated by newlines
        and columns separated by tabs.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to read Excel files. Install with `pip install openpyxl`."
        )

    text = []
    workbook = load_workbook(path)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            # Filter out None values and convert to strings
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text.append(row_text)

    return "\n".join(text)


def read_all_documents(folder: str = "documents") -> Dict[str, str]:
    """Read all supported documents from a folder.

    Supported file types:
    - .txt
    - .pdf
    - .docx
    - .doc
    - .xlsx
    - .xls

    Args:
        folder: Folder path containing documents.

    Returns:
        A dictionary mapping each file name to its extracted text content.
    """
    documents: Dict[str, str] = {}
    folder_path = Path(folder)

    """
    'iterdir()' iterates through the immediate contents of that directory.
    It returns both files and subdirectories.
    It does not search recursively into subfolders.
    """

    for file_path in sorted(folder_path.iterdir()):
        if not file_path.is_file():
            continue

        suffix = file_path.suffix.lower()
        if suffix == ".txt":
            documents[file_path.name] = read_text_file(str(file_path))
        elif suffix == ".pdf":
            documents[file_path.name] = read_pdf_file(str(file_path))
        elif suffix in (".docx", ".doc"):
            documents[file_path.name] = read_docx_file(str(file_path))
        elif suffix in (".xlsx", ".xls"):
            documents[file_path.name] = read_excel_file(str(file_path))

    return documents

# ------------------------------------------------------------------
# Test Section
# ------------------------------------------------------------------
# The code inside this block executes only when this file is run
# directly from the command line.
#
# Example:
#     python3 app/extract_pdf.py
#
# In this case:
#     __name__ == "__main__"  → True
#
# Therefore, the test code below will execute.
#
# However, when this file is imported into another module:
#
#     from extract_pdf import read_all_documents
#
# Then:
#     __name__ == "extract_pdf"   → False
#
# As a result, the test code is skipped and only the functions
# defined in this file are imported.
#
# This is a standard Python practice used to:
#   • Keep reusable functions separate from test code
#   • Prevent test code from running during imports
#   • Allow the file to be executed independently for debugging
# ------------------------------------------------------------------

if __name__ == "__main__":
    all_docs = read_all_documents()
    print(f"✓ Loaded {len(all_docs)} document(s):")
    for name, text in all_docs.items():
        print(f"  - {name} ({len(text)} chars)")
